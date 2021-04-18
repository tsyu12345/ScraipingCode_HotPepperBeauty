from openpyxl.worksheet.dimensions import SheetDimension
from requests.api import options
from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, InvalidSwitchToTargetException, NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys
from openpyxl import styles as pxstyle
import openpyxl as px 
from bs4 import BeautifulSoup as bs 
import time
import datetime
import re
import requests as rq

class Job:

    def __init__(self, driver_path, books_path):
        self.book_path = books_path
        self.book = px.load_workbook(books_path)
        self.sheet = self.book.worksheets[0]
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-gpu')
        self.driver = webdriver.Chrome(executable_path=driver_path)
        self.driver.set_window_size(1128, 768)
    
    def url_scrap(self, area_name, store_class):
        print("starting ChromeDriver.exe....")
        self.driver.get("https://beauty.hotpepper.jp/top/")#top page
        sr_class = self.driver.find_element_by_link_text(store_class)
        sr_class.click()
        time.sleep(1)
        search = self.driver.find_element_by_css_selector('#freeWordSearch1')
        search.send_keys(area_name + Keys.ENTER)
        time.sleep(2)
        result_count = self.driver.find_element_by_css_selector('#mainContents > div.mT20.bgWhite > div.preListHead > div > p:nth-child(1) > span').text 
        result_pages = self.driver.find_element_by_css_selector('#mainContents > div.mT20.bgWhite > div.preListHead > div > p.pa.bottom0.right0').text
        page_num = re.split('[/ ]', result_pages)
        pages = re.sub(r"\D", "", page_num[1])
        print("Search Result : %s : %s" % (area_name, result_count))
        print("pages : " + pages)
        link_list = []
        for i in range(int(pages)):
            html = self.driver.page_source
            soup = bs(html, 'lxml')
            links_list = soup.select("div.slcHeadContentsInner > h3 > a")
            for a in links_list:
                url = a.get('href')
                link_list.append(url)
            try:
                next_btn = self.driver.find_element_by_link_text("次へ")
                next_btn.click()
                time.sleep(1)
            except NoSuchElementException:
                break
        print("search complete")
     
        for i in range(len(link_list)):
            self.sheet.cell(row=i+2, column=1, value=store_class)
            self.sheet.cell(row=i+2, column=8, value=link_list[i])
        self.book.save(self.book_path)
        

    def info_scrap(self, url, index):
        dt_now = datetime.datetime.now()
        year = str(dt_now.year)
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        data_day = year + "," + month + day + "," + hour + min
        print(data_day)
        self.driver.get(url)
        print(url)
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        store_name_tag = soup.select_one('p.detailTitle > a')
        store_name = store_name_tag.get_text()
        print("店名：" + store_name)
        st_name_kana_tag = soup.select_one('#mainContents > div.detailHeader.cFix.pr > div > div.pL10.oh.hMin120 > div > p.fs10.fgGray')
        st_name_kana = st_name_kana_tag.get_text()
        print("店名カナ：" + st_name_kana)
        try:
            tel_tag = soup.select_one('div.mT30 > table > tbody > tr > td > a')
            tel_url = tel_tag.get('href')
            respons_tel = rq.get(tel_url)
            html_tel = respons_tel.text
            soup_tel = bs(html_tel, 'lxml')
            tel_num_tag = soup_tel.select_one('table > tr > td')
            tel_num = tel_num_tag.get_text()
            print("TEL : " + tel_num)
            self.sheet.cell(row=index, column=4, value=tel_num)
        except:
            pass     
        
        #ヘッダー画像の有無
        try:
            self.driver.find_element_by_css_selector('#jsiNavCarousel > div')
            self.sheet.cell(row=index, column=14, value="有")
        except NoSuchElementException:
            self.sheet.cell(row=index, column=14, value="無")
            pass

        table_value = soup.select('div.mT30 > table > tbody > tr > td')
        table_menu = soup.select('div.mT30 > table > tbody > tr > th')
        print(table_menu)
        #住所の抽出（少々処理があるため別で書き出す）
        for i, e in enumerate(table_menu):
            if e.get_text() == "住所":    
                all_address = table_value[i].get_text()
                prefecture_search = re.search('東京都|北海道|(?:京都|大阪)府|.{2,3}県' , all_address)
                address_low = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)#県名とそれ以降を分離
                prefecture = prefecture_search.group()#県名
                jis_code = self.call_jis_code(prefecture)
                municipality = address_low[1]#それ以降
        print("都道府県：" + prefecture)
        print("市区町村番地：" + municipality)
    
        #他の情報処理
        for i in range(2, len(table_value)):
            for c in range(1, self.sheet.max_column):
                if table_menu[i].get_text() == self.sheet.cell(row=1, column=c).value:
                    self.sheet.cell(row=index, column=c, value=table_value[i].get_text())
                    break
                
        """
        anounce_access = table[2].get_text()
        bs_time = table[3].get_text()
        holiday = table[4].get_text()
        credit = table[5].get_text()
        fee = table[6].get_text()
        seat_cnt = table[7].get_text()
        staff_cnt = table[8].get_text()
        parking = table[9].get_text()
        commitment = table[10].get_text()
        try:
            remarks = table[11].get_text()
        except IndexError:
            remarks = ""
"""
        catch_copy_tag = soup.select_one('#mainContents > div.pH10.mT25 > div:nth-child(1) > p > b > strong')
        catch_copy = catch_copy_tag.get_text()

        pankuzu_tag = soup.select('#preContents > ol > li')
        pankuzu = ""
        for pan in pankuzu_tag:
            pankuzu += pan.get_text()
        print(pankuzu)

        slide_img_tag = soup.select('#mainContents > div.pH10.mT25 > div:nth-child(1) > div > div.slnTopImg.jscThumbCarousel > div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
        slide_cnt = len(slide_img_tag)
 
        self.sheet.cell(row=index, column=2, value=store_name)
        self.sheet.cell(row=index, column=3, value=st_name_kana)
        #self.sheet.cell(row=index, column=4, value=tel_num)
        self.sheet.cell(row=index, column=5, value=jis_code)
        self.sheet.cell(row=index, column=6, value=prefecture)
        self.sheet.cell(row=index, column=7, value=municipality)
        self.sheet.cell(row=index, column=9, value=data_day)
        self.sheet.cell(row=index, column=13, value=pankuzu)
        self.sheet.cell(row=index, column=16, value=slide_cnt)
        self.sheet.cell(row=index, column=17, value=catch_copy)
        """
        self.sheet.cell(row=index, column=18, value=anounce_access)
        self.sheet.cell(row=index, column=19, value=bs_time)
        self.sheet.cell(row=index, column=20, value=holiday)
        self.sheet.cell(row=index, column=21, value=credit)
        self.sheet.cell(row=index, column=23, value=fee)
        self.sheet.cell(row=index, column=24, value=seat_cnt)
        self.sheet.cell(row=index, column=25, value=staff_cnt)
        self.sheet.cell(row=index, column=26, value=parking)
        self.sheet.cell(row=index, column=27, value=commitment)
        self.sheet.cell(row=index, column=28, value=remarks)
        """
        self.book.save(self.book_path)

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        print(code)
        return code

    def apper_adjust(self, index):
        for c in range(1, self.sheet.max_column+1):
            if self.sheet.cell(row=index, column=c).value == None:
                self.sheet.cell(row=index, column=c, value=" ")
        self.book.save(self.book_path)

    def check_prefecture(self, area_name, index):
        for i in range(2, self.sheet.max_row+1):
            if self.sheet.cell(row=index, column=6).value != area_name:
                for c in range(1, self.sheet.max_column+1):
                    self.sheet.cell(row=index, column=c, value=None)
        self.book.save(self.book_path)

job = Job("chromedriver_win32\chromedriver.exe", "【サンプル】ホットペッパービューティー copy.xlsx")
job.url_scrap("北海道", "ヘアサロン")
for i in range(2, job.sheet.max_row+1):
    try:
        if job.sheet.cell(row=i, column=6).value == None:
            print(job.sheet.cell(row=i, column=8).value)
            job.info_scrap(job.sheet.cell(row=i, column=8).value, i)
            job.apper_adjust(index=i)
    except WebDriverException:
        print("WebDriver Exception occured. Retrying...")
        job.driver.close()
        job.info_scrap(job.sheet.cell(row=i, column=8).value, i)
        job.apper_adjust(index=i)

print("==WARNING==")
print("checking scrap data.!!DO NOT STOP THIS PROGRAM!!.")
print("if you close this program the loading file will be break.")
print("checking.")    
for i in range(2, job.sheet.max_row+1):
    print(".", end="")
    job.check_prefecture('北海道', i)
print("complete data check. you can open the '.xlsx' file.")




        

       