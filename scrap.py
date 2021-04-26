from openpyxl.worksheet.dimensions import SheetDimension
from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, InvalidSwitchToTargetException, NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys
from openpyxl import styles as pxstyle
import openpyxl as px
from bs4 import BeautifulSoup as bs
import time
import datetime
import re
import requests as rq
import concurrent.futures as cf


class Job:

    def __init__(self, driver_path, books_path, area_name, store_class):
        self.book_path = books_path
        self.book = px.load_workbook(books_path)
        self.sheet = self.book.worksheets[0]
        """
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-gpu')
        """
        self.driver = webdriver.Chrome(executable_path=driver_path)
        self.driver.set_window_size(1128, 768)
        self.area_name = area_name
        self.store_class = store_class

    def url_scrap(self):
        print("starting ChromeDriver.exe....")
        self.driver.get("https://beauty.hotpepper.jp/top/")  # top page
        sr_class = self.driver.find_element_by_link_text(self.store_class)
        sr_class.click()
        time.sleep(1)
        search = self.driver.find_element_by_css_selector('#freeWordSearch1')
        search.send_keys(self.area_name + Keys.ENTER)
        time.sleep(2)
        result_count = self.driver.find_element_by_css_selector(
            '#mainContents > div.mT20.bgWhite > div.preListHead > div > p:nth-child(1) > span').text
        result_pages = self.driver.find_element_by_css_selector(
            '#mainContents > div.mT20.bgWhite > div.preListHead > div > p.pa.bottom0.right0').text
        page_num = re.split('[/ ]', result_pages)
        pages = re.sub(r"\D", "", page_num[1])
        print("Search Result : %s : %s" % (self.area_name, result_count))
        print("pages : " + pages)
        link_list = []
        for i in range(int(pages)):
            html = self.driver.page_source
            soup = bs(html, 'lxml')
            links_list = soup.select("div.slcHeadContentsInner > h3 > a")
            for a in links_list:
                url = a.get('href')
                link_list.append(url)
            try:
                next_btn = self.driver.find_element_by_link_text("次へ")
                next_btn.click()
                time.sleep(1)
            except NoSuchElementException:
                break
        print("search complete")

        for i in range(len(link_list)):
            self.sheet.cell(row=i+2, column=1, value=self.store_class)
            self.sheet.cell(row=i+2, column=8, value=link_list[i])
        self.book.save(self.book_path)

    def info_scrap(self, url, index):
        dt_now = datetime.datetime.now()
        year = str(dt_now.year)
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        data_day = year + "," + month + day + "," + hour + min
        print(data_day)
        self.driver.get(url)
        print(url)
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        store_name_tag = soup.select_one('p.detailTitle > a')
        store_name = store_name_tag.get_text()
        print("店名：" + store_name)
        st_name_kana_tag = soup.select_one(
            '#mainContents > div.detailHeader.cFix.pr > div > div.pL10.oh.hMin120 > div > p.fs10.fgGray')
        st_name_kana = st_name_kana_tag.get_text()
        print("店名カナ：" + st_name_kana)
        try:
            tel_tag = soup.select_one('div.mT30 > table > tbody > tr > td > a')
            tel_url = tel_tag.get('href')
            respons_tel = rq.get(tel_url)
            html_tel = respons_tel.text
            soup_tel = bs(html_tel, 'lxml')
            tel_num_tag = soup_tel.select_one('table > tr > td')
            tel_num = tel_num_tag.get_text()
            print("TEL : " + tel_num)
        except:
            pass

        # ヘッダー画像の有無
        try:
            self.driver.find_element_by_css_selector('#jsiNavCarousel > div')
            head_img_yn = "有"
            # self.sheet.cell(row=index, column=14, value="有")
        except NoSuchElementException:
            head_img_yn = "無"
            # self.sheet.cell(row=index, column=14, value="無")
            pass

        table_value = soup.select('div.mT30 > table > tbody > tr > td')
        table_menu = soup.select('div.mT30 > table > tbody > tr > th')
        print(table_menu)
        # 住所の抽出（少々処理があるため別で書き出す）
        for i, e in enumerate(table_menu):
            if e.get_text() == "住所":
                all_address = table_value[i].get_text()
                prefecture_search = re.search(
                    '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                address_low = re.split(
                    '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                prefecture = prefecture_search.group()  # 県名
                if prefecture != self.area_name:
                    self.sheet.cell(row=index, column=8, value="")
                    return False
                jis_code = self.call_jis_code(prefecture)
                municipality = address_low[1]  # それ以降
        print("都道府県：" + prefecture)
        print("市区町村番地：" + municipality)

        catch_copy_tag = soup.select_one(
            '#mainContents > div.pH10.mT25 > div:nth-child(1) > p > b > strong')
        catch_copy = catch_copy_tag.get_text()

        pankuzu_tag = soup.select('#preContents > ol > li')
        pankuzu = ""
        for pan in pankuzu_tag:
            pankuzu += pan.get_text()
        print(pankuzu)

        slide_img_tag = soup.select(
            '#mainContents > div.pH10.mT25 > div:nth-child(1) > div > div.slnTopImg.jscThumbCarousel > div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
        slide_cnt = len(slide_img_tag)

# write Excel
        self.sheet.cell(row=index, column=2, value=store_name)
        self.sheet.cell(row=index, column=3, value=st_name_kana)
        self.sheet.cell(row=index, column=4, value=tel_num)
        self.sheet.cell(row=index, column=5, value=jis_code)
        self.sheet.cell(row=index, column=6, value=prefecture)
        self.sheet.cell(row=index, column=7, value=municipality)
        self.sheet.cell(row=index, column=9, value=data_day)
        self.sheet.cell(row=index, column=13, value=pankuzu)
        self.sheet.cell(row=index, column=16, value=slide_cnt)
        self.sheet.cell(row=index, column=17, value=catch_copy)
        self.sheet.cell(row=index, column=14, value=head_img_yn)
        # 他の情報処理
        for i in range(2, len(table_value)):
            for c in range(1, self.sheet.max_column):
                if table_menu[i].get_text() == self.sheet.cell(row=1, column=c).value:
                    self.sheet.cell(row=index, column=c,
                                    value=table_value[i].get_text())
                    break
        self.book.save(self.book_path)

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        print(code)
        return code

    def apper_adjust(self, index):
        for c in range(1, self.sheet.max_column+1):
            if self.sheet.cell(row=index, column=c).value == None:
                self.sheet.cell(row=index, column=c, value=" ")
        self.book.save(self.book_path)

    # 指定行までのURLリストを生成し返す
    def make_url_list(self, stop_index):
        url_list = []
        for i in range(2, stop_index):
            url_list.append(self.sheet.cell(row=i, column=8).value)
        return url_list


if __name__ == "__main__":
    job = Job("chromedriver_win32\chromedriver.exe",
              "sample1 - コピー.xlsx", "北海道", "ヘアサロン")

    def process(s_inedx, end_index, process_number):
        
        print("start process NO." + process_number)
        for i in range(s_inedx, end_index+1):
            if job.sheet.cell(row=i, column=8).value == None:
                print("runnning index" + str(i))
                job.url_scrap()
        for i in range(s_inedx, end_index+1):
            try:
                if job.sheet.cell(row=i, column=6).value == None:
                    print("runnning scrap_method index" + str(i))
                    print(job.sheet.cell(row=i, column=8).value)
                    job.info_scrap(job.sheet.cell(row=i, column=8).value, i)
                    job.apper_adjust(index=i)
            except WebDriverException:
                print("WebDriver Exception occured. Retrying...")
                job.driver.close()
                job.info_scrap(job.sheet.cell(row=i, column=8).value, i)
                job.apper_adjust(index=i)
        print("end process No." + process_number)

    exxuter = cf.ThreadPoolExecutor(max_workers=4)
    exxuter.submit(process, 2, 100, "1")
    exxuter.submit(process, 101, 200, "2")

""" 
    for i in range(2, job.sheet.max_row+1):
        if job.sheet.cell(row=i, column=8).value == None:
            job.url_scrap()
    for i in range(2, job.sheet.max_row+1):
        try:
            if job.sheet.cell(row=i, column=6).value == None:
                print(job.sheet.cell(row=i, column=8).value)
                job.info_scrap(job.sheet.cell(row=i, column=8).value, i)
                job.apper_adjust(index=i)
        except WebDriverException:
            print("WebDriver Exception occured. Retrying...")
            job.driver.close()
            job.info_scrap(job.sheet.cell(row=i, column=8).value, i)
            job.apper_adjust(index=i)
"""
"""
    print("==WARNING==")
    print("checking scrap data.!!DO NOT STOP THIS PROGRAM!!.")
    print("if you close this program the loading file will be break.")
    print("checking.")    
    for i in range(2, job.sheet.max_row+1):
        print(".", end="")
        job.check_prefecture('北海道', i)
    print("complete data check. you can open the '.xlsx' file.")
    end_time = time.time() - st_time
    print("{0}".format(end_time) + "[sec]")
"""
