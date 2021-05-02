from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, InvalidSwitchToTargetException, NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys
from openpyxl import cell, styles as pxstyle
import openpyxl as px
from bs4 import BeautifulSoup as bs
import time
import datetime
import re
import requests as rq
import threading as th


class Job:
    """
    book_path = 'sample1-原本 - コピー.xlsx'
    book = px.load_workbook(book_path)
    sheet = book.worksheets[0]
    """
    book = px.Workbook()
    sheet = book.worksheets[0]
    def init(self, driver_path, save_path, area, store_class):
        #init driver
        self.driver_path = driver_path
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('--headless')
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-gpu')
        self.area_name = area
        self.store_class = store_class
        #init workBook
        """
        n_book = px.Workbook()
        n_sheet = n_book.worksheets[0]
        """
        font = pxstyle.Font(name='游ゴシック')
        menu = [
            "ジャンル",
            "店舗名",
            "店舗名カナ",
            "電話番号", 	
            "都道府県コード",
            "都道府県", 
            "市区町村・番地",
            "店舗URL",
            "詳細データ取得日",
            "料金プラン着地",
            "月額料金",	
            "お店のホームページ",
            "パンくず",	
            "ヘッダー画像有無",	
            "こだわり有無",
            "スライド画像数",
            "キャッチコピー",
            "アクセス・道案内",
            "営業時間",
            "定休日",
            "支払い方法",
            "設備",
            "カット価格",
            "席数",	
            "スタッフ数",
            "駐車場",	
            "こだわり条件",
            "備考",
            "スタッフ募集",
            "最終更新日"				
        ]
        for c in range(1, 30+1):
            self.sheet.cell(row=1, column=c, value=menu[c-1])
        self.sheet.freeze_panes = "A2"
        """
        for r in self.sheet:
            for c in r:
                self.sheet[].font = font
        """
        dt_now = datetime.datetime.now()
        month = str(dt_now.month)
        day = str(dt_now.day)
        name = month + day + "_result.xlsx"
        self.book_path = save_path + name
        self.book.save(save_path + name)
        """
        self.book_path = 'sample1-原本 - コピー.xlsx'
        self.book = px.load_workbook(self.book_path)
        self.sheet = self.book.worksheets[0]
        """
    

    
    def url_scrap(self):
        print("starting ChromeDriver.exe....")
        driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        driver.get("https://beauty.hotpepper.jp/top/")  # top page
        sr_class = driver.find_element_by_link_text(self.store_class)
        sr_class.click()
        time.sleep(1)
        search = driver.find_element_by_css_selector('#freeWordSearch1')
        search.send_keys(self.area_name + Keys.ENTER)
        time.sleep(2)
        result_count = driver.find_element_by_css_selector(
            '#mainContents > div.mT20.bgWhite > div.preListHead > div > p:nth-child(1) > span').text
        result_pages = driver.find_element_by_css_selector(
            '#mainContents > div.mT20.bgWhite > div.preListHead > div > p.pa.bottom0.right0').text
        page_num = re.split('[/ ]', result_pages)
        pages = re.sub(r"\D", "", page_num[1])
        print("Search Result : %s : %s" % (self.area_name, result_count))
        print("pages : " + pages)
        link_list = []
        for i in range(int(pages)):
            html = driver.page_source
            soup = bs(html, 'lxml')
            links_list = soup.select("div.slcHeadContentsInner > h3 > a")
            for a in links_list:
                url = a.get('href')
                print(url)
                link_list.append(url)
            try:
                next_btn = driver.find_element_by_link_text("次へ")
                next_btn.click()
                time.sleep(1)
            except NoSuchElementException:
                break
        print("search complete")

        for i in range(len(link_list)):
            self.sheet.cell(row=i+2, column=1, value=self.store_class)
            self.sheet.cell(row=i+2, column=8, value=link_list[i])
        self.book.save(self.book_path)
        driver.quit()
        
    def info_scrap(self, start_index, end_index):
        driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        for i in range(start_index, end_index+1):
            driver.get(self.sheet.cell(row=i, column=8).value)
            html = driver.page_source
            soup = bs(html, 'lxml')
            store_name_tag = soup.select_one('p.detailTitle > a')
            store_name = store_name_tag.get_text()
            print("店名：" + store_name)
            st_name_kana_tag = soup.select_one(
                '#mainContents > div.detailHeader.cFix.pr > div > div.pL10.oh.hMin120 > div > p.fs10.fgGray')
            st_name_kana = st_name_kana_tag.get_text()
            print("店名カナ：" + st_name_kana)
            try:
                tel_tag = soup.select_one('div.mT30 > table > tbody > tr > td > a')
                tel_url = tel_tag.get('href')
                respons_tel = rq.get(tel_url)
                html_tel = respons_tel.text
                soup_tel = bs(html_tel, 'lxml')
                tel_num_tag = soup_tel.select_one('table > tr > td')
                tel_num = tel_num_tag.get_text()
                print("TEL : " + tel_num)
            except:
                tel_num = ""
                pass
            # ヘッダー画像の有無
            try:
                driver.find_element_by_css_selector('#jsiNavCarousel > div')
                head_img_yn = "有"
                # self.sheet.cell(row=index, column=14, value="有")
            except NoSuchElementException:
                head_img_yn = "無"
                # self.sheet.cell(row=index, column=14, value="無")
                pass
            table_value = soup.select('div.mT30 > table > tbody > tr > td')
            table_menu = soup.select('div.mT30 > table > tbody > tr > th')
            print(table_menu)
            print(table_value)
            # 住所の抽出（少々処理があるため別で書き出す）
            pref_tf = True
            for j, e in enumerate(table_menu):
                if e.get_text() == "住所":
                    all_address = table_value[j].get_text()
                    prefecture_search = re.search(
                        '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                    address_low = re.split(
                        '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                    prefecture = prefecture_search.group()  # 県名
                    if prefecture != self.area_name:
                        self.sheet.cell(row=i, column=8, value="")
                        self.sheet.cell(row=i, column=1, value="")
                        pref_tf = False                        
                        break
                    jis_code = self.call_jis_code(prefecture)
                    municipality = address_low[1]  # それ以降
            
            print("都道府県：" + prefecture)
            print("市区町村番地：" + municipality)
            catch_copy_tag = soup.select_one('#mainContents > div.pH10.mT25 > div:nth-child(1) > p > b > strong')
            catch_copy = catch_copy_tag.get_text()

            pankuzu_tag = soup.select('#preContents > ol > li')
            pankuzu = ""
            for pan in pankuzu_tag:
                pankuzu += pan.get_text()
            print(pankuzu)

            slide_img_tag = soup.select('#mainContents > div.pH10.mT25 > div:nth-child(1) > div > div.slnTopImg.jscThumbCarousel > div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
            slide_cnt = len(slide_img_tag)

            # write Excel
            self.sheet.cell(row=i, column=2, value=store_name)
            self.sheet.cell(row=i, column=3, value=st_name_kana)
            self.sheet.cell(row=i, column=4, value=tel_num)
            self.sheet.cell(row=i, column=5, value=jis_code)
            self.sheet.cell(row=i, column=6, value=prefecture)
            self.sheet.cell(row=i, column=7, value=municipality)
            self.sheet.cell(row=i, column=9, value=self.scrap_day())
            self.sheet.cell(row=i, column=13, value=pankuzu)
            self.sheet.cell(row=i, column=16, value=slide_cnt)
            self.sheet.cell(row=i, column=17, value=catch_copy)
            self.sheet.cell(row=i, column=14, value=head_img_yn)
            # 他の情報処理
            try:
                for j in range(2, len(table_value)):
                    for c in range(1, self.sheet.max_column):
                        if table_menu[j].get_text() == self.sheet.cell(row=1, column=c).value:
                            self.sheet.cell(row=i, column=c, value=table_value[j].get_text())
                            break
            except RuntimeError:
                pass
            #self.book_save()
    
    def book_save(self):
        self.book.save(self.book_path)

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        print(code)
        return code

    def scrap_day(self):
        dt_now = datetime.datetime.now()
        year = str(dt_now.year)
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        data_day = year + "," + month + day + "," + hour + min
        print(data_day)
        return data_day

def range_segmentation(maxindex):
    rang = list(range(2, maxindex, int(maxindex / 5)))
    print(rang)
    return rang

def process0(path, area, st_class):
    job = Job()
    job.init('chromedriver_win32\chromedriver.exe', path, area, st_class)
    job.url_scrap()
    job.book_save()

def process1(s_index, e_index, path, area, st_class):
    job = Job()
    job.init('chromedriver_win32\chromedriver.exe', path, area, st_class)
    job.info_scrap(s_index, e_index)
    #job.book.save()

def load_max_row():
    job = Job()
    max_row = job.sheet.max_row
    return max_row

def multiThread(path, area, st_class):
        sheet_index = load_max_row()
        rng = range_segmentation(sheet_index)
        th_s = []
        for i in range(4):
            if i == 0:
                add = th.Thread(target=process1, args=(rng[0], rng[1], path, area, st_class))
            else:
                add = th.Thread(target=process1, args=(rng[i]+1, rng[i+1], path, area, st_class))
            th_s.append(add)
        add = th.Thread(target=process1, args=(rng[4]+1, sheet_index, path, area, st_class))
        th_s.append(add)

        for i in range(len(th_s)):
            th_s[i].start()
        for i in range(len(th_s)):
            th_s[i].join()

def main(path, area, st_class):
    process0(path, area, st_class)
    multiThread(path, area, st_class)
"""
area_name = gui.value['pref_name']
st_class  = gui.value['store_class']
save_path = gui.value['path']
process0(save_path, area_name, st_class)
multiThread(save_path, area_name, st_class)
"""
if __name__ == "__main__":
    
    def singleThread():
        process1(2, 100)
    
    def multiThread():
        job = Job()
        sheet_index = load_max_row()
        rng = range_segmentation(sheet_index)
        th_s = []
        for i in range(4):
            if i == 0:
                add = th.Thread(target=process1, args=(rng[0], rng[1]))
            else:
                add = th.Thread(target=process1, args=(rng[i]+1, rng[i+1]))
            th_s.append(add)
        add = th.Thread(target=process1, args=(rng[4]+1, sheet_index))
        th_s.append(add)

        for i in range(len(th_s)):
            th_s[i].start()
        for i in range(len(th_s)):
            th_s[i].join()
        job.book_save()
    """
    s_time_single = time.time()
    singleThread()
    end_time_single = time.time() - s_time_single
    print("{0}".format(end_time_single) + "[sec]")
    """
    s_time = time.time()
    process0()
    multiThread()
    end_time = time.time() - s_time
    print("{0}".format(end_time) + "[sec]")
    